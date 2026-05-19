import streamlit as st
import pickle
import numpy as np
import os
import io
import tempfile
import shutil

import openpyxl
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

# ── CONFIG ────────────────────────────────────────────────────────────────────

INDEX_PATH = "cache/png_index.pkl"

# ── Model & data ──────────────────────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def load_model():
    return SentenceTransformer("all-MiniLM-L6-v2")  # 👈 local saved model


@st.cache_data(show_spinner=False)
def load_data():
    with open("chunked_data.pkl", "rb") as f:  # 👈 chunked data
        docs = pickle.load(f)
    embs = np.load("embeddings.npy")
    return docs, embs


@st.cache_data(show_spinner=False)
def load_png_index() -> dict:
    if not os.path.exists(INDEX_PATH):
        return {}
    with open(INDEX_PATH, "rb") as f:
        raw_index = pickle.load(f)
    basename_index = {}
    for (file_path, sheet_name), png_path in raw_index.items():
        base_key = (os.path.basename(file_path), sheet_name)
        basename_index[base_key] = png_path
    return basename_index


def get_png_path(png_index: dict, file_path: str, sheet_name: str) -> str | None:
    key      = (os.path.basename(file_path), sheet_name)
    png_path = png_index.get(key)
    if png_path and os.path.exists(png_path):
        return png_path
    return None


def extract_sheet_as_xlsx(file_path: str, sheet_name: str) -> bytes | None:
    try:
        # Load the source workbook
        src_wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name not in src_wb.sheetnames:
            # Try strip matching
            matched = next(
                (s for s in src_wb.sheetnames if s.strip() == sheet_name.strip()),
                None
            )
            if matched is None:
                return None
            sheet_name = matched

        src_ws = src_wb[sheet_name]

        # Create a new workbook and copy the sheet data
        dst_wb = openpyxl.Workbook()
        dst_ws = dst_wb.active
        dst_ws.title = sheet_name

        for row in src_ws.iter_rows(values_only=True):
            dst_ws.append(list(row))

        # Save to bytes
        output = io.BytesIO()
        dst_wb.save(output)
        output.seek(0)
        return output.read()

    except Exception as e:
        st.caption(f"Sheet extraction error: {e}")
        return None


# ── UI ────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Excel Semantic Search", layout="wide")
st.title("📊 Excel Semantic Search")

model                    = load_model()
chunked_data, embeddings = load_data()  # 👈 updated
png_index                = load_png_index()

if not png_index:
    st.warning(
        "⚠️ PNG cache not found. "
        "Run `python generate_png_cache.py` first, then restart the app."
    )

top_k = 3

query = st.text_input(
    "🔍 Enter your query",
    placeholder="e.g. quarterly revenue by region",
)

if query:
    query_embedding = model.encode([query])
    scores          = cosine_similarity(query_embedding, embeddings)[0]

    query_lower = query.lower().strip()

    for i, chunk in enumerate(chunked_data):
        chunk_lower = chunk['text'].lower()

        # Highest boost - clause is the main subject
        if f"clause no.: {query_lower}" in chunk_lower:
            scores[i] += 1.0  # 👈 very strong boost

        # Medium boost - clause mentioned directly
        elif query_lower in chunk_lower:
            scores[i] += 0.3

        # Small boost - partial keyword match
        else:
            query_keywords = query_lower.split()
            matches = sum(1 for kw in query_keywords if kw in chunk_lower)
            scores[i] += 0.05 * matches

    top_indices = scores.argsort()[-top_k:][::-1]

    st.subheader("Relevant Sheets")

    for rank, idx in enumerate(top_indices, 1):

        doc        = chunked_data[idx]  # 👈 updated
        meta       = doc["metadata"]
        file_name  = meta["file_name"]
        sheet_name = meta["sheet_name"]
        score      = round(float(scores[idx]), 3)

        with st.expander(
            f"#{rank} · {os.path.basename(file_name)} — sheet: {sheet_name} (score: {score})",
            expanded=(rank == 1),
        ):

            col1, col2 = st.columns([3, 1])

            with col1:
                st.caption(
                    f"File: {file_name} | Sheet: {sheet_name} | Score: {score}"
                )

            with col2:
                sheet_bytes = extract_sheet_as_xlsx(file_name, sheet_name)

                if sheet_bytes:
                    download_name = f"{os.path.splitext(os.path.basename(file_name))[0]} — {sheet_name}.xlsx"
                    st.download_button(
                        label="⬇ Download Sheet",
                        data=sheet_bytes,
                        file_name=download_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{idx}",
                    )
                else:
                    try:
                        with open(file_name, "rb") as fh:
                            data = fh.read()
                        ext  = os.path.splitext(file_name)[1].lower()
                        mime = (
                            "application/vnd.ms-excel"
                            if ext == ".xls"
                            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.download_button(
                            label="⬇ Download Excel",
                            data=data,
                            file_name=os.path.basename(file_name),
                            mime=mime,
                            key=f"dl_{idx}",
                        )
                    except FileNotFoundError:
                        st.warning("File not found for download.")

            st.markdown("**Text preview:**")
            st.code(doc["text"][:300], language=None)  # 👈 updated

            st.markdown("**Sheet preview:**")

            png_path = get_png_path(png_index, file_name, sheet_name)

            if png_path:
                st.image(png_path, caption=sheet_name, use_container_width=True)
            else:
                st.info(f"Preview not available for: {sheet_name}")